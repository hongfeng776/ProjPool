import csv
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from calculation import MechanicalCalculator

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class BatchCalculator:
    def __init__(self, max_workers=4):
        self.calculator = MechanicalCalculator()
        self.batch_data = []
        self.batch_results = []
        self.max_workers = max_workers

    def load_from_csv(self, file_path):
        if not os.path.exists(file_path):
            return False, "文件不存在"
        
        try:
            self.batch_data = []
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    data = self._parse_row(row)
                    if data:
                        self.batch_data.append(data)
            
            if len(self.batch_data) == 0:
                return False, "CSV文件中没有有效数据"
            
            return True, f"成功加载 {len(self.batch_data)} 条数据"
        except Exception as e:
            return False, f"加载失败: {str(e)}"

    def _parse_row(self, row):
        try:
            data = {
                "part_id": row.get("零件编号", "").strip(),
                "part_name": row.get("零件名称", "").strip(),
                "part_type": row.get("零件类型", "").strip(),
                "material": row.get("材料", "Q235").strip(),
                "length": self._parse_float(row.get("长度")),
                "width": self._parse_float(row.get("宽度")),
                "height": self._parse_float(row.get("高度")),
                "force": self._parse_float(row.get("受力")),
                "force_direction": row.get("受力方向", "").strip()
            }
            return data
        except Exception:
            return None

    def _parse_float(self, value):
        if value is None or value.strip() == "":
            return None
        try:
            return float(value.strip())
        except ValueError:
            return None

    def add_single_data(self, data):
        self.batch_data.append(data)

    def clear_data(self):
        self.batch_data = []
        self.batch_results = []

    def _calculate_single(self, item):
        idx, data = item
        calc = MechanicalCalculator()
        material = data.get("material", "Q235")
        result = calc.calculate_all(data, material=material)
        return {
            "index": idx + 1,
            "part_id": data.get("part_id", ""),
            "part_name": data.get("part_name", ""),
            "part_type": data.get("part_type", ""),
            "input_data": data,
            **result
        }

    def calculate_all(self, use_parallel=True):
        start_time = time.time()
        self.batch_results = []
        
        if use_parallel and len(self.batch_data) > 10:
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {executor.submit(self._calculate_single, (i, data)): i 
                          for i, data in enumerate(self.batch_data)}
                for future in as_completed(futures):
                    self.batch_results.append(future.result())
            self.batch_results.sort(key=lambda x: x["index"])
        else:
            for i, data in enumerate(self.batch_data):
                self.batch_results.append(self._calculate_single((i, data)))
        
        elapsed = time.time() - start_time
        return self.batch_results, elapsed

    def get_results_summary(self):
        if not self.batch_results:
            return "暂无计算结果"
        
        summary = []
        summary.append(f"批量计算完成，共 {len(self.batch_results)} 条记录")
        summary.append("-" * 100)
        summary.append(f"{'序号':<6}{'零件名称':<15}{'材料':<10}{'截面积':<12}{'拉应力':<12}{'安全系数':<10}{'状态':<10}")
        summary.append("-" * 100)
        
        for r in self.batch_results:
            results = r.get("results", {})
            status = "✓" if r.get("success") else "✗"
            summary.append(
                f"{r['index']:<6}"
                f"{r['part_name']:<15}"
                f"{results.get('material', '-'):<10}"
                f"{results.get('cross_sectional_area', '-'):<12}"
                f"{results.get('tensile_stress', '-'):<12}"
                f"{results.get('safety_factor', '-'):<10}"
                f"{status:<10}"
            )
        
        return "\n".join(summary)

    def export_results_to_csv(self, output_path):
        if not self.batch_results:
            return False, "没有可导出的结果"
        
        try:
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow([
                    "序号", "零件编号", "零件名称", "零件类型", "材料",
                    "长度(mm)", "宽度(mm)", "高度(mm)", "受力(N)", "受力方向",
                    "截面积(mm²)", "拉伸应力(MPa)", "压缩应力(MPa)", "剪切应力(MPa)",
                    "轴向应变", "伸长量(mm)", "安全系数", "计算状态", "警告信息"
                ])
                
                for r in self.batch_results:
                    data = r["input_data"]
                    results = r.get("results", {})
                    warnings = "; ".join(r.get("warnings", [])) if r.get("warnings") else ""
                    status = "成功" if r.get("success") else "失败"
                    
                    writer.writerow([
                        r["index"],
                        r["part_id"],
                        r["part_name"],
                        r["part_type"],
                        results.get("material", ""),
                        data.get("length", ""),
                        data.get("width", ""),
                        data.get("height", ""),
                        data.get("force", ""),
                        data.get("force_direction", ""),
                        results.get("cross_sectional_area", ""),
                        results.get("tensile_stress", ""),
                        results.get("compressive_stress", ""),
                        results.get("shear_stress", ""),
                        results.get("axial_strain", ""),
                        results.get("elongation", ""),
                        results.get("safety_factor", ""),
                        status,
                        warnings
                    ])
            
            return True, f"结果已导出到: {output_path}"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

    def export_report_to_excel(self, output_path):
        if not EXCEL_AVAILABLE:
            return False, "未安装 openpyxl 库，无法导出Excel格式"
        
        if not self.batch_results:
            return False, "没有可导出的结果"
        
        try:
            wb = openpyxl.Workbook()
            
            ws_summary = wb.active
            ws_summary.title = "计算汇总"
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            ws_summary.merge_cells('A1:L1')
            title_cell = ws_summary['A1']
            title_cell.value = "零部件力学计算报告"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            ws_summary['A3'] = "计算时间:"
            ws_summary['B3'] = time.strftime("%Y-%m-%d %H:%M:%S")
            ws_summary['A4'] = "计算数量:"
            ws_summary['B4'] = f"{len(self.batch_results)} 件"
            
            success_count = sum(1 for r in self.batch_results if r.get("success"))
            ws_summary['A5'] = "成功数量:"
            ws_summary['B5'] = f"{success_count} 件"
            ws_summary['A6'] = "失败数量:"
            ws_summary['B6'] = f"{len(self.batch_results) - success_count} 件"
            
            headers = ["序号", "零件编号", "零件名称", "零件类型", "材料", 
                      "截面积(mm²)", "拉应力(MPa)", "压应力(MPa)", 
                      "安全系数", "状态", "警告数"]
            
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=8, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            for row_idx, r in enumerate(self.batch_results, 9):
                results = r.get("results", {})
                status = "成功" if r.get("success") else "失败"
                warning_count = len(r.get("warnings", []))
                
                values = [
                    r["index"],
                    r["part_id"],
                    r["part_name"],
                    r["part_type"],
                    results.get("material", ""),
                    results.get("cross_sectional_area", ""),
                    results.get("tensile_stress", ""),
                    results.get("compressive_stress", ""),
                    results.get("safety_factor", ""),
                    status,
                    warning_count
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws_summary.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    if not r.get("success"):
                        cell.fill = error_fill
                    elif warning_count > 0:
                        cell.fill = warning_fill
            
            for col in range(1, len(headers) + 1):
                ws_summary.column_dimensions[chr(64 + col)].width = 14
            
            ws_detail = wb.create_sheet("详细数据")
            
            detail_headers = [
                "序号", "零件编号", "零件名称", "零件类型", "材料",
                "长度(mm)", "宽度(mm)", "高度(mm)", "受力(N)", "受力方向",
                "截面积(mm²)", "拉伸应力(MPa)", "压缩应力(MPa)", "剪切应力(MPa)",
                "轴向应变", "伸长量(mm)", "安全系数", "弹性模量(GPa)", "屈服强度(MPa)",
                "计算状态", "警告信息", "错误信息"
            ]
            
            for col, header in enumerate(detail_headers, 1):
                cell = ws_detail.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            for row_idx, r in enumerate(self.batch_results, 2):
                data = r["input_data"]
                results = r.get("results", {})
                warnings = "; ".join(r.get("warnings", [])) if r.get("warnings") else ""
                errors = "; ".join(r.get("errors", [])) if r.get("errors") else ""
                status = "成功" if r.get("success") else "失败"
                
                values = [
                    r["index"],
                    r["part_id"],
                    r["part_name"],
                    r["part_type"],
                    results.get("material", ""),
                    data.get("length", ""),
                    data.get("width", ""),
                    data.get("height", ""),
                    data.get("force", ""),
                    data.get("force_direction", ""),
                    results.get("cross_sectional_area", ""),
                    results.get("tensile_stress", ""),
                    results.get("compressive_stress", ""),
                    results.get("shear_stress", ""),
                    results.get("axial_strain", ""),
                    results.get("elongation", ""),
                    results.get("safety_factor", ""),
                    results.get("elastic_modulus", ""),
                    results.get("yield_strength", ""),
                    status,
                    warnings,
                    errors
                ]
                
                for col, value in enumerate(values, 1):
                    cell = ws_detail.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    if not r.get("success"):
                        cell.fill = error_fill
                    elif r.get("warnings"):
                        cell.fill = warning_fill
            
            for col in range(1, len(detail_headers) + 1):
                ws_detail.column_dimensions[chr(64 + col)].width = 14
            
            wb.save(output_path)
            return True, f"Excel报告已导出到: {output_path}"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

    def generate_sample_csv(self, output_path):
        try:
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["零件编号", "零件名称", "零件类型", "材料", "长度", "宽度", "高度", "受力", "受力方向"])
                writer.writerow(["P001", "齿轮A", "齿轮", "Q235", 100, 50, 20, 50000, "径向"])
                writer.writerow(["P002", "轴B", "轴类", "45钢", 200, 30, 30, 80000, "轴向"])
                writer.writerow(["P003", "螺栓C", "螺栓", "40Cr", 80, 10, 10, 20000, "轴向"])
                writer.writerow(["P004", "弹簧D", "弹簧", "Q345", 150, 25, 25, 30000, "轴向"])
                writer.writerow(["P005", "支架E", "板类", "不锈钢304", 120, 60, 15, 45000, "横向"])
            
            return True, f"示例CSV已生成: {output_path}"
        except Exception as e:
            return False, f"生成失败: {str(e)}"
